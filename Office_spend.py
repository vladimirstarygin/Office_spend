#!/usr/bin/env python
# coding: utf-8

# In[12]:


from dialog_bot_sdk.bot import DialogBot
from dialog_bot_sdk import interactive_media

import openpyxl
import grpc

file = 'data.xlsx'
is_adding_period = [False for i in range(5)]
is_adding_cons = [False for i in range(5)]
is_out_cons = False
is_delete_period = False
is_delete_con = [False for i in range(3)]
data_start = [" " for i in range(3)]
data_end = [" " for i in range(3)]
total_cons = "0"
data_cons = [" " for i in range(4)]
ws_delete = ""
peer = ""

def on_msg(*params):
    global peer
    peer = params[0].peer
    message = params[0].message.textMessage.text
    message = message.lower()
    if adding_period(message, peer):
        print("")
    elif adding_cons(message, peer):
        print("")
    elif is_out_cons:
        out_cons(message, peer)
    elif is_delete_con[2]:
        delete_con(message, peer)
    elif is_delete_period:
        delete_period(message, peer)
    elif message == 'help':
        help(message, peer)
    elif message == 'добавить период':
        add_period(peer)
    elif message == 'добавить расходы':
        add_cons(peer)
    elif message == 'вывести расходы':
        out_cons(message, peer)
    elif message == 'удалить расход':
        delete_con(message, peer)
    elif message == 'удалить период':
        delete_period(message, peer)
    elif message == 'скачать файл':
        get_file(peer)
    else:
        bot.messaging.send_message(peer, "Введите 'help' для помощи")

# In[2]:

def add_period(peer):
    if not is_adding_period[4]:
        msg = "Введите дату начала периода в формате '1 января 2019' "
        msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
        btn = interactive_media.InteractiveMedia(1,
        interactive_media.InteractiveMediaButton("отмена", "отмена"))
        bot.messaging.send_message(
            peer,
            msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        is_adding_period[0] = True
        is_adding_period[4] = True
        
def adding_period(message, peer):
    global data_start, data_end, total_cons, is_adding_period
    if(is_adding_period[0]):
        data = message.split()
        if (len(data) == 3):
            if RepresentInt(data[0]) and RepresentMonth(data[1]) and RepresentInt(data[2]):
                day = int(data[0])
                month = data[1]
                year = data[2]
                if day > 0 and day <= 31:
                    msg = "Введите дату окончания периода в том же формате"
                    msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
                    btn = interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("отмена", "отмена"))
                    bot.messaging.send_message(
                        peer,
                        msg, [interactive_media.InteractiveMediaGroup([ btn ])])
                    is_adding_period[0] = False
                    is_adding_period[1] = True
                    data_start = data
                else:
                    msg = "Введите правильное число"
                    msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
                    btn = interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("отмена", "отмена"))
                    bot.messaging.send_message(
                        peer,
                        msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            else:
                msg = "Введите правильную дату"
                msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
                btn = interactive_media.InteractiveMedia(1,
                interactive_media.InteractiveMediaButton("отмена", "отмена"))
                bot.messaging.send_message(
                    peer,
                    msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        else:
            msg = "Введите правильную дату"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        return True
    elif(is_adding_period[1]):
        data = message.split()
        if (len(data) == 3):
            if RepresentInt(data[0]) and RepresentMonth(data[1]) and RepresentInt(data[2]):
                day = int(data[0])
                month = data[1]
                year = data[2]
                if day > 0 and day <= 31:
                    msg = "Введите суммарные расходы для периода в формате '20000'"
                    msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
                    btn = interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("отмена", "отмена"))
                    bot.messaging.send_message(
                        peer,
                        msg, [interactive_media.InteractiveMediaGroup([ btn ])])
                    is_adding_period[1] = False
                    is_adding_period[2] = True
                    data_end = data
                else:
                    msg = "Введите правильное число"
                    msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
                    btn = interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("отмена", "отмена"))
                    bot.messaging.send_message(
                        peer,
                        msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            else:
                msg = "Введите правильную дату"
                msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
                btn = interactive_media.InteractiveMedia(1,
                interactive_media.InteractiveMediaButton("отмена", "отмена"))
                bot.messaging.send_message(
                    peer,
                    msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        else:
            msg = "Введите правильную дату"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        return True
    elif(is_adding_period[2]):
        if RepresentFloat(message):
            total_cons = message.replace(",",".")
            msg = "Введите название расходного периода"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            is_adding_period[2] = False
            is_adding_period[3] = True
        else:
            msg = "Введите число"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        return True
    elif(is_adding_period[3]):
        wb = openpyxl.load_workbook(filename = file)
        sheets = wb.sheetnames
        if message not in sheets:
            wb.create_sheet(message)
            ws = wb[message]
            ws['A1'] = "Начало периода:"
            ws['B1'] = data_start[0]
            ws['C1'] = data_start[1]
            ws['D1'] = data_start[2]
            ws['A2'] = "Окончание периода:"
            ws['B2'] = data_end[0]
            ws['C2'] = data_end[1]
            ws['D2'] = data_end[2]
            ws['E1'] = "Суммарные расходы:"
            ws['F1'] = total_cons
            ws['E2'] = "Остаток:"
            ws['F2'] = total_cons
            ws['A3'] = "номер"
            ws['B3'] = "наименование"
            ws['C3'] = "количество"
            ws['D3'] = "стоимость"
            msg = "Cтатья расходов '"+ message + "' с суммарными расходами " + total_cons 
            msg += " с датой начала " + " ".join(data_start) + " года и датой окончания "
            msg += " ".join(data_end) + " года добавлена! "
            bot.messaging.send_message(peer, msg)
            is_adding_period = [False for i in range(5)]
            data_start = [" " for i in range(3)]
            data_end = [" " for i in range(3)]
            total_cons = "0"
            wb.save(file)
        else:
            msg = "Статья расходов с таким названием уже есть, введите другое"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления периода"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        return True
    else:
        return False


# In[3]:


def add_cons(peer):
    if not is_adding_cons[4]:
        wb = openpyxl.load_workbook(filename = file)
        sheets = wb.sheetnames
        if len(sheets) == 1:
            msg = "Добавьте хотя бы один расходный период"
            bot.messaging.send_message(peer, msg)
        else:
            msg = "Выберите расходный период:\n"
            msg += "\n".join(sheets[1:len(sheets)])
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления расходов"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            is_adding_cons[0] = True
            is_adding_cons[4] = True
        
def adding_cons(message, peer):
    global data_cons, is_adding_cons
    if(is_adding_cons[0]):
        wb = openpyxl.load_workbook(filename = file)
        sheets = wb.sheetnames
        if message not in sheets[1:len(sheets)]:
            msg = "Выберите один из представленных расходных периодов:\n"
            msg += "\n".join(sheets[1:len(sheets)])
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления расходов"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            return True
        else:
            data_cons[0] = message
            msg = "Введите наименование расхода:"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления расходов"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            is_adding_cons[0] = False
            is_adding_cons[1] = True
            return True
    if(is_adding_cons[1]):
        data_cons[1] = message
        msg = "Введите количество расходуемых товаров:"
        msg += "\nили нажмите кнопку 'отмена' для завершения добавления расходов"
        btn = interactive_media.InteractiveMedia(1,
        interactive_media.InteractiveMediaButton("отмена", "отмена"))
        bot.messaging.send_message(
            peer,
            msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        is_adding_cons[1] = False
        is_adding_cons[2] = True
        return True
    if is_adding_cons[2]:
        if not RepresentInt(message):
            msg = "Введите число"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления расходов"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            return True
        else:
            data_cons[2] = message
            msg = "Введите суммарную стоимость всех товаров:"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления расходов"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            is_adding_cons[2] = False
            is_adding_cons[3] = True
            return True
    if is_adding_cons[3]:
        if not RepresentFloat(message):
            msg = "Введите число"
            msg += "\nили нажмите кнопку 'отмена' для завершения добавления расходов"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            return True
        else:
            data_cons[3] = message.replace(",",".")
            wb = openpyxl.load_workbook(filename = file)
            ws = wb[data_cons[0]]
            pos = len(ws['A'])
            if RepresentInt(ws['A'+str(pos)].value):
                count = int(ws['A'+str(pos)].value) + 1
            else:
                count = 1
            ws['A'+str(pos+1)] = count
            ws['B'+str(pos+1)] = data_cons[1]
            ws['C'+str(pos+1)] = data_cons[2]
            ws['D'+str(pos+1)] = data_cons[3]
            
            balance = float(str(ws['F2'].value).replace(",","."))
            balance -= float(data_cons[3])
            ws['F2'] = balance
            msg = "Расход № "+str(count)+":\n '"
            msg += " ".join(data_cons[1:len(data_cons)])+"' "
            msg += "добавлен в таблицу расходов '"+data_cons[0]+"'\n"
            msg += "остаток: " + str(balance)
            bot.messaging.send_message(peer, msg)
            wb.save(file)
            is_adding_cons = [False for i in range(5)]
            data_cons = [" " for i in range(4)]
            return True
        
    else:
        return False


# In[4]:


def out_cons(message, peer):
    global is_out_cons, file
    wb = openpyxl.load_workbook(filename = file)
    sheets = wb.sheetnames
    if not is_out_cons:
        msg = "Выберите расходный период:\n"
        msg += "\n".join(sheets[1:len(sheets)])
        msg += "\nили нажмите кнопку 'отмена' для завершения отображения всех расходов"
        btn = interactive_media.InteractiveMedia(1,
        interactive_media.InteractiveMediaButton("отмена", "отмена"))
        bot.messaging.send_message(
            peer,
            msg, [interactive_media.InteractiveMediaGroup([ btn ])])
    else:
        if message in sheets:
            ws = wb[message]
            msg = "" + str(ws['A1'].value) + " "
            msg += str(ws['B1'].value) + " "
            msg += str(ws['C1'].value) + " "
            msg += str(ws['D1'].value) + "\n"
            msg += str(ws['A2'].value) + " "
            msg += str(ws['B2'].value) + " "
            msg += str(ws['C2'].value) + " "
            msg += str(ws['D2'].value) + "\n"
            msg += str(ws['E1'].value) + " "
            msg += str(ws['F1'].value) + "\n"
            msg += str(ws['E2'].value) + " "
            msg += str(ws['F2'].value) + "\n"
            pos = len(ws['A'])
            for i in range(4,pos+1):
                msg += str(ws['A'+str(i)].value) + ". "
                msg += str(ws['B'+str(i)].value) + " "
                msg += str(ws['C'+str(i)].value) + " "
                msg += str(ws['D'+str(i)].value) + "\n"
            bot.messaging.send_message(peer, msg)
            is_out_cons = False
            return
        else:
            msg = "Выберите расходный период:\n"
            msg += "\n".join(sheets[1:len(sheets)])
            msg += "\nили нажмите кнопку 'отмена' для завершения отображения всех расходов"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
    is_out_cons = True


# In[5]:


def delete_con(message, peer):
    global is_delete_con, file, ws_delete
    wb = openpyxl.load_workbook(filename = file)
    sheets = wb.sheetnames
    if not is_delete_con[2]:
        msg = "Выберите расходный период:\n"
        msg += "\n".join(sheets[1:len(sheets)])
        msg += "\nили нажмите кнопку 'отмена' для завершения удаления расхода"
        btn = interactive_media.InteractiveMedia(1,
        interactive_media.InteractiveMediaButton("отмена", "отмена"))
        bot.messaging.send_message(
            peer,
            msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        is_delete_con[2] = True
        is_delete_con[0] = True
    else:
        if is_delete_con[0]:
            if message in sheets[1:len(sheets)]:
                ws = wb[message]
                ws_delete = message
                pos = len(ws['A'])
                if pos == 3:
                    msg = "В расходном периоде нет ни одного расхода, выберите другой."
                    msg += "\nили нажмите кнопку 'отмена' для завершения удаления расхода"
                    btn = interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("отмена", "отмена"))
                    bot.messaging.send_message(
                        peer,
                        msg, [interactive_media.InteractiveMediaGroup([ btn ])])
                else:
                    msg = "Выберете номер расхода для удаления:\n"
                    for i in range(4,pos+1):
                        msg += str(ws['A'+str(i)].value) + ". "
                        msg += str(ws['B'+str(i)].value) + " "
                        msg += str(ws['C'+str(i)].value) + " "
                        msg += str(ws['D'+str(i)].value) + "\n"
                    msg += "\nили нажмите кнопку 'отмена' для завершения удаления расхода"
                    btn = interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("отмена", "отмена"))
                    bot.messaging.send_message(
                        peer,
                        msg, [interactive_media.InteractiveMediaGroup([ btn ])])
                    is_delete_con[0] = False
                    is_delete_con[1] = True
            else:
                msg = "Выберите расходный период:\n"
                msg += "\n".join(sheets[1:len(sheets)])
                msg += "\nили нажмите кнопку 'отмена' для завершения удаления расхода"
                btn = interactive_media.InteractiveMedia(1,
                interactive_media.InteractiveMediaButton("отмена", "отмена"))
                bot.messaging.send_message(
                    peer,
                    msg, [interactive_media.InteractiveMediaGroup([ btn ])])
        elif is_delete_con[1]:
            ws = wb[ws_delete]
            pos = len(ws['A'])
            if RepresentInt(message):
                delete = int(message)
                if delete >= int(ws['A4'].value) and delete <= int(ws['A'+str(pos)].value):
                    ws['F2'] = str(float(ws['F2'].value) + float(ws['D'+str(3+delete)].value))
                    msg = "Расход '"
                    msg += str(ws['A'+str(3+delete)].value) + ". "
                    msg += str(ws['B'+str(3+delete)].value) + " "
                    msg += str(ws['C'+str(3+delete)].value) + " "
                    msg += str(ws['D'+str(3+delete)].value) + "' "
                    for i in range(3+delete+1,pos+1):
                        ws['A'+str(i-1)].value = str(int(ws['A'+str(i)].value)-1)
                        ws['B'+str(i-1)].value = ws['B'+str(i)].value
                        ws['C'+str(i-1)].value = ws['C'+str(i)].value
                        ws['D'+str(i-1)].value = ws['D'+str(i)].value
                    ws['A'+str(pos)] = None
                    ws['B'+str(pos)] = None
                    ws['C'+str(pos)] = None
                    ws['D'+str(pos)] = None
                    wb.save(file)
                    msg += "удален"
                    bot.messaging.send_message(peer, msg)
                    is_delete_con = [False for i in range(3)]
                else:
                    msg = "Выберите число в диапазоне от 1" + " до " + str(ws['A'+str(pos)].value)
                    msg += "\nили нажмите кнопку 'отмена' для завершения удаления расхода"
                    btn = interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("отмена", "отмена"))
                    bot.messaging.send_message(
                        peer,
                        msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            else:
                msg = "Введите число в диапазоне от 1" + " до " + str(ws['A'+str(pos)].value)
                msg += "\nили нажмите кнопку 'отмена' для завершения удаления расхода"
                btn = interactive_media.InteractiveMedia(1,
                interactive_media.InteractiveMediaButton("отмена", "отмена"))
                bot.messaging.send_message(
                    peer,
                    msg, [interactive_media.InteractiveMediaGroup([ btn ])])


# In[6]:


def RepresentInt(s):
    try: 
        number = int(s)
        if number < 0:
            return False
        else:
            return True
    except ValueError:
        return False
    
def RepresentFloat(s):
    try: 
        number = float(s.replace(",","."))
        if number < 0 or number != number :
            return False
        else:
            return True
    except ValueError:
        return False

def RepresentMonth(s):
    month = ['декабрь','декабря','январь','января','февраль','февраля']
    month += ['март','марта','апрель','апреля','май', 'мая']
    month += ['июнь','июня','июль','июля','август', 'августа']
    month += ['сентябрь','сентября','октябрь','октября','ноябрь', 'ноября']
    if s in month:
        return True
    else:
        return False


# In[7]:


def delete_period(message, peer):
    global is_delete_period, file
    if not is_delete_period:
        wb = openpyxl.load_workbook(filename = file)
        sheets = wb.sheetnames
        if len(sheets) > 1:
            msg = "Выберите расходный период для удаления:\n"
            msg += "\n".join(sheets[1:len(sheets)])
            msg += "\nили нажмите кнопку 'отмена' для завершения удаления периода"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])
            is_delete_period = True
        else:
            msg = "Добавтье хотя бы один расчетный период"
            bot.messaging.send_message(
                peer,
                msg)
    else:
        wb = openpyxl.load_workbook(filename = file)
        sheets = wb.sheetnames
        if message in sheets:
            wb.remove(wb[message])
            wb.save(file)
            msg = "Расчетный период " + message + " удален."
            is_delete_period = False
            bot.messaging.send_message(peer, msg)
        else:
            msg = "Выберите расходный период для удаления:\n"
            msg += "\n".join(sheets[1:len(sheets)])
            msg += "\nили нажмите кнопку 'отмена' для завершения удаления периода"
            btn = interactive_media.InteractiveMedia(1,
            interactive_media.InteractiveMediaButton("отмена", "отмена"))
            bot.messaging.send_message(
                peer,
                msg, [interactive_media.InteractiveMediaGroup([ btn ])])


# In[8]:


def get_file(peer):
    global file
    msg = "Держи!"
    bot.messaging.send_message(peer, msg)
    bot.messaging.send_file(peer, file)


# In[9]:


def help(message, peer):
    msg  = "Введите или выберите:\n"
    msg += "'добавить период' для добавления периода с остатком денежных средств;\n"
    msg += "'добавить расходы' для добавления расходов в выбранном периоде;\n"
    msg += "'вывести расходы' для отображения расходов для выбранного периода;\n"
    msg += "'удалить расход' для удаления расхода для выбранного периода;\n"
    msg += "'удалить период' для удаления периода со всеми расходами;\n"
    msg += "'скачать файл' для скачивания Excel файла со всеми расходами." 
    bot.messaging.send_message(
        peer,
        msg,
        [interactive_media.InteractiveMediaGroup(
            [
                interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("добавить период", "добавить период")),
                interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("добавить расходы", "добавить расходы")),
                interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("вывести расходы", "вывести расходы")),
                interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("удалить расход", "удалить расход")),
                interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("удалить период", "удалить период")),
                interactive_media.InteractiveMedia(1,
                    interactive_media.InteractiveMediaButton("скачать файл", "скачать файл")),
            ]
        )]
    )


# In[10]:


def on_click(*params):
    global peer
    which_button = params[0].value
    uid = params[0].uid
    message = ""
    peer = bot.users.get_user_peer_by_id(uid)
    if which_button == 'добавить период':
        reload()
        add_period(peer)
    elif which_button == 'добавить расходы':
        reload()
        add_cons(peer)
    elif which_button == 'вывести расходы':
        reload()
        out_cons(message, peer)
    elif which_button == 'удалить расход':
        reload()
        delete_con(message, peer)
    elif which_button == 'удалить период':
        reload()
        delete_period(message, peer)
    elif which_button == 'скачать файл':
        reload()
        get_file(peer)
    elif which_button == "отмена":
        reload()


# In[11]:


def reload():
    global is_adding_period, is_adding_cons, is_out_cons, is_delete_period
    global is_delete_con, data_start, data_end, total_cons, data_cons, ws_delete
    global peer
    
    if is_adding_period[4]:
        msg = "Добавление расчетного периода отменено"
        bot.messaging.send_message(peer, msg)
    elif is_adding_cons[4]:
        msg = "Добавление расхода отменено"
        bot.messaging.send_message(peer, msg)
    elif is_out_cons:
        msg = "Отображение всех расходов для периода отменено"
        bot.messaging.send_message(peer, msg)
    elif is_delete_period:
        msg = "Удаление периода отменено"
        bot.messaging.send_message(peer, msg)
    elif is_delete_con[2]:
        msg = "Удаление расхода отменено"
        bot.messaging.send_message(peer, msg)
    is_adding_period = [False for i in range(5)]
    is_adding_cons = [False for i in range(5)]
    is_out_cons = False
    is_delete_period = False
    is_delete_con = [False for i in range(3)]
    data_start = [" " for i in range(3)]
    data_end = [" " for i in range(3)]
    total_cons = "0"
    data_cons = [" " for i in range(4)]
    ws_delete = ""


if __name__ == '__main__':
    bot = DialogBot.get_secure_bot(
        "hackathon-mob.transmit.im",  # bot endpoint from environment
        grpc.ssl_channel_credentials(), # SSL credentials (empty by default!)
        '7f38e457bebadf578dbba6d5ccbacd4b3e926938'  # bot token from environment
    )
    bot.messaging.on_message(on_msg, on_click)
